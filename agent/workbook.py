"""Write the four submission workbooks by filling the supplied templates.

We load challenge/templates/<FILE>.xlsx, locate the Summary-sheet header row
(Metric | Units | <period>) exactly the way scripts/check-forecasts.mjs does, and
write the forecast number into the period column for each metric row — touching
nothing else, so the Summary structure the checker validates stays byte-identical.
"""
from __future__ import annotations

import os

from openpyxl import load_workbook

from .corpus import ROOT

TEMPLATES = os.path.join(ROOT, "challenge", "templates")
SUBMISSION = os.path.join(ROOT, "submission")


def _round(value: float, kind: str) -> float:
    if kind == "eps":
        return round(value, 2)
    if kind == "pct":
        return round(value, 2)
    return round(value, 1)          # money (USDm / GBPm)


def write_direct(output_file: str, period: str, metrics: list[dict]) -> str:
    """Write from selected metric dicts (nested or direct fallback): each dict has
    label, kind and point. Fills the Summary-sheet period column, structure untouched."""
    src = os.path.join(TEMPLATES, output_file)
    wb = load_workbook(src)
    ws = wb["Summary"]
    header = None
    for r in range(1, 31):
        if (str(ws.cell(r, 1).value or "").strip() == "Metric"
                and str(ws.cell(r, 2).value or "").strip() == "Units"
                and str(ws.cell(r, 3).value or "").strip() == period):
            header = r
            break
    if header is None:
        raise RuntimeError(f"{output_file}: Metric/Units/{period} header not found")
    by_label = {m["label"]: m for m in metrics}
    for i in range(1, 20):
        r = header + i
        label = str(ws.cell(r, 1).value or "").strip()
        if not label:
            break
        m = by_label.get(label)
        if m and m.get("point") is not None:
            ws.cell(r, 3).value = _round(m["point"], m.get("kind") or "money")
    os.makedirs(SUBMISSION, exist_ok=True)
    out = os.path.join(SUBMISSION, output_file)
    wb.save(out)
    return out


def write_workbook(fc: dict) -> str:
    """fc = forecast_company(...) result. Returns the written path."""
    src = os.path.join(TEMPLATES, fc["output_file"])
    wb = load_workbook(src)
    ws = wb["Summary"]

    # find header row: col1 "Metric", col2 "Units", col3 == period
    header = None
    for r in range(1, 31):
        if (str(ws.cell(r, 1).value or "").strip() == "Metric"
                and str(ws.cell(r, 2).value or "").strip() == "Units"
                and str(ws.cell(r, 3).value or "").strip() == fc["period"]):
            header = r
            break
    if header is None:
        raise RuntimeError(f"{fc['output_file']}: Metric/Units/{fc['period']} header not found")

    by_label = {m.label: m for m in fc["metrics"]}
    for i in range(1, 20):
        r = header + i
        label = str(ws.cell(r, 1).value or "").strip()
        if not label:
            break
        m = by_label.get(label)
        if m and m.point is not None:
            ws.cell(r, 3).value = _round(m.point, m.kind)

    os.makedirs(SUBMISSION, exist_ok=True)
    out = os.path.join(SUBMISSION, fc["output_file"])
    wb.save(out)
    return out
